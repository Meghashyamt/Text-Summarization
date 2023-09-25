from flask import Flask, render_template, request
import logging
from transformers import AutoTokenizer, AutoModelForSeq2SeqLM, pipeline
from summarizer import Summarizer
from PyPDF2 import PdfReader
import re
import pandas as pd
import spacy
import time,os
import openpyxl
from openpyxl import load_workbook


# Load spaCy language model
nlp = spacy.load("en_core_web_sm")

# Load a pretrained summarization model
model_name = "facebook/bart-large-cnn"
tokenizer = AutoTokenizer.from_pretrained(model_name)
model = AutoModelForSeq2SeqLM.from_pretrained(model_name)


app = Flask(__name__)

df = pd.DataFrame(columns=['Keyword', 'Extracted Content', 'Hugging Face Summary', 'Summarization Library Summary'])

@app.route("/", methods=["GET", "POST"])
def index():
    global df
    if request.method == "POST":
        uploaded_file = request.files["file"]
        keyword = request.form["keyword"]

        # Inside your `index` route function
        if uploaded_file and keyword:
            # Measure the start time for content extraction
            extraction_start_time = time.time()

            # Save the uploaded file
            uploaded_file.save("uploaded_file.pdf")

            # Extract text from the PDF document
            document = extract_text_from_pdf("uploaded_file.pdf")

            # Measure the end time for content extraction
            extraction_end_time = time.time()

            # Calculate the time taken for content extraction
            extraction_time = extraction_end_time - extraction_start_time

            # Count words in the extracted content
            word_count = len(document.split())

            # Extract content around the keyword
            relevant_content = extract_content_around_keyword(document, keyword)

            # Measure the start time for summarization
            summarization_start_time = time.time()

            # Summarize the extracted content using Hugging Face model
            if relevant_content:
                relevant_inputs = tokenizer(relevant_content, return_tensors="pt", max_length=1024, truncation=True)
                summary_ids = model.generate(relevant_inputs["input_ids"], max_length=1500, min_length=40, length_penalty=2.0, num_beams=4, early_stopping=True)
                summary_hf = tokenizer.decode(summary_ids[0], skip_special_tokens=True)
            else:
                summary_hf = "No relevant content found."

            # Measure the end time for summarization
            summarization_end_time = time.time()

            # Calculate the time taken for summarization
            summarization_time = summarization_end_time - summarization_start_time

            # Summarize the extracted content using the 'summarizer' library
            summarizer_model = Summarizer()
            summarizer_start_time = time.time()
            summarizer_summary = summarizer_model(relevant_content)
            summarizer_end_time = time.time()
            summarizer_time = summarizer_end_time - summarizer_start_time

            summarizer = pipeline("summarization", model="google/pegasus-xsum")
            #summary = summarizer(relevant_content, min_length=30, do_sample=False)
            import torch
            #from transformers import AutoTokenizer, AutoModelForSeq2SeqLM
            model1 = AutoModelForSeq2SeqLM.from_pretrained('t5-base')
            tokenizer1 = AutoTokenizer.from_pretrained('t5-base')

            tokens_input = tokenizer1.encode("summarize: " + relevant_content,
                                        return_tensors='pt',
                                        max_length=tokenizer1.model_max_length,
                                        truncation=True)

            summary_ids = model1.generate(tokens_input, min_length=80, 
                                    max_length=150, length_penalty=15, 
                                    num_beams=2)
            summary = tokenizer1.decode(summary_ids[0], skip_special_tokens=True)
            #print(summary)
            
            new_data = {'Keyword': keyword, 'Extracted Content': relevant_content, 'Hugging Face Summary': summary_hf, 'Summarization Library Summary': summary}
            df = df.append(new_data, ignore_index=True)

            # Save the DataFrame to an Excel file
            #df.to_excel(r'C:\Users\M_Thiruveedula\Downloads\Solution_python\Solution_python\python_solution\summaries.xlsx', index=False)
            excel_file_path = r'C:\Users\M_Thiruveedula\Downloads\Solution_python\Solution_python\python_solution\summaries.xlsx'

            # Check if the Excel file exists, create it if not
            # ...

            # Check if the Excel file exists
            if not os.path.isfile(excel_file_path):
                # If the file doesn't exist, create it with the DataFrame
                df.to_excel(excel_file_path, index=False, sheet_name='Sheet1')
            else:
                # Load the existing Excel file
                book = load_workbook(excel_file_path)
                
                # Create a Pandas Excel writer object with the loaded workbook
                writer = pd.ExcelWriter(excel_file_path, engine='openpyxl') 
                writer.book = book

                # Check if 'Sheet1' exists, create it if not (this should be done only once)
                if 'Sheet1' not in writer.book.sheetnames:
                    writer.book.create_sheet(title='Sheet1')

                # Get the last row number in 'Sheet1'
                sheet = writer.book['Sheet1']
                last_row = sheet.max_row if sheet.max_row is not None else 0

                # Append the DataFrame to the Excel file without overwriting
                df.to_excel(writer, index=False, sheet_name='Sheet1', startrow=last_row + 1)

                # Save the Excel file
                writer.save()
                writer.close()

            # ...



            return render_template("index.html", keyword=keyword, extracted_content=relevant_content, hf_summary=summary_hf,
                                summarizer_summary=summary, spacy_summary=summarizer, extraction_time=extraction_time,
                                word_count=word_count, summarization_time=summarization_time, summarizer_time=summarizer_time)


    return render_template("index.html", keyword="", extracted_content="", hf_summary="", summarizer_summary="")

#if __name__ == "__main__":
#    app.run(debug=True)
def extract_text_from_pdf(pdf_path):
    text = ""
    pdf_reader = PdfReader(pdf_path)
    for page in pdf_reader.pages:
        text += page.extract_text()
    return text

def extract_content_around_keyword(document, keyword, max_words=300):
    keyword_positions = [match.start() for match in re.finditer(re.escape(keyword), document, re.IGNORECASE)]
    extracted_content = ""

    for position in keyword_positions:
        start = max(0, position - max_words)
        end = min(len(document), position + max_words)
        extracted_content += document[start:end]
    
    extracted_content = re.sub(r"[^a-zA-Z0-9\s]", "", extracted_content)

    extracted_content = re.sub(r"\d+", "", extracted_content)
    extracted_content = extracted_content.replace("Section", "")

    return extracted_content

if __name__ == "__main__":
    app.run(debug=True)
